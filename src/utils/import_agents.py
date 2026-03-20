import json
import re
from datetime import datetime
from openpyxl import load_workbook
from core.database import Database
from core.models import Agent


def parse_date(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    s = str(value).strip()
    if s.upper() in ('N/A', '', 'NONE'):
        return None
    return s


def parse_split(split_str: str | None, notes: str | None) -> dict:
    if split_str is None:
        split_str = ''
    split_str = str(split_str).strip()
    notes = str(notes).strip() if notes else ''

    # Transaction fee agents: split is "0" and notes mention a fee
    fee_match = re.search(r'\$(\d+)\s*transaction fee', notes, re.IGNORECASE)
    if fee_match:
        fee = float(fee_match.group(1))
        return {
            'split_type': 'transaction_fee',
            'agent_split_pct': None,
            'office_split_pct': None,
            'tier_rules': None,
            'transaction_fee': fee,
        }

    # Tiered: "60/40 First 4 then 70/30"
    tiered_match = re.match(
        r'(\d+)/(\d+)\s*[Ff]irst\s*(\d+)\s*then\s*(\d+)/(\d+)',
        split_str
    )
    if tiered_match:
        a1, o1, count, a2, o2 = tiered_match.groups()
        tier_rules = json.dumps({
            'tiers': [
                {'max_txn_count': int(count), 'agent_pct': int(a1), 'office_pct': int(o1)},
                {'max_txn_count': None, 'agent_pct': int(a2), 'office_pct': int(o2)},
            ]
        })
        return {
            'split_type': 'tiered',
            'agent_split_pct': float(a1),
            'office_split_pct': float(o1),
            'tier_rules': tier_rules,
            'transaction_fee': None,
        }

    # Standard percentage: "70/30", "80/20", etc.
    pct_match = re.match(r'(\d+)/(\d+)', split_str)
    if pct_match:
        agent_pct = float(pct_match.group(1))
        office_pct = float(pct_match.group(2))
        return {
            'split_type': 'percentage',
            'agent_split_pct': agent_pct,
            'office_split_pct': office_pct,
            'tier_rules': None,
            'transaction_fee': None,
        }

    # N/A, blank, or "0" without a transaction fee note -> inactive
    return {
        'split_type': 'percentage',
        'agent_split_pct': None,
        'office_split_pct': None,
        'tier_rules': None,
        'transaction_fee': None,
    }


def import_from_excel(db: Database, excel_path: str) -> int:
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    imported = 0
    for row_idx in range(4, 33):  # rows 4-32 (1-indexed)
        name = ws.cell(row=row_idx, column=1).value
        if not name or str(name).strip() == '':
            continue
        name = str(name).strip()

        license_num = ws.cell(row=row_idx, column=2).value
        license_num = str(license_num).strip() if license_num else None

        license_exp = parse_date(ws.cell(row=row_idx, column=3).value)

        split_str = ws.cell(row=row_idx, column=4).value
        cap_val = ws.cell(row=row_idx, column=5).value
        contract_date = parse_date(ws.cell(row=row_idx, column=6).value)
        w9_date = parse_date(ws.cell(row=row_idx, column=7).value)
        notes = ws.cell(row=row_idx, column=8).value
        notes = str(notes).strip() if notes else ''

        # Parse cap amount
        cap_amount = None
        if cap_val is not None:
            cap_str = str(cap_val).strip()
            if cap_str.upper() not in ('N/A', '', 'NONE'):
                try:
                    cap_amount = float(cap_str)
                except ValueError:
                    cap_amount = None

        # Parse split info
        split_info = parse_split(str(split_str) if split_str is not None else '', notes)

        # Determine if active
        is_active = 1
        split_str_check = str(split_str).strip() if split_str else ''
        if split_str_check.upper() in ('N/A', '') and split_info['split_type'] != 'transaction_fee':
            is_active = 0
        if 'not currently employed' in notes.lower():
            is_active = 0
        # Agents with blank split and no fee info are inactive
        if split_info['agent_split_pct'] is None and split_info['transaction_fee'] is None:
            is_active = 0

        agent = Agent(
            id=0,
            name=name,
            license_number=license_num,
            license_expiration=license_exp,
            split_type=split_info['split_type'],
            agent_split_pct=split_info['agent_split_pct'],
            office_split_pct=split_info['office_split_pct'],
            tier_rules=split_info['tier_rules'],
            transaction_fee=split_info['transaction_fee'],
            cap_amount=cap_amount,
            contract_date=contract_date,
            is_active=is_active,
            notes=notes,
        )
        db.insert_agent(agent)
        imported += 1

    wb.close()
    return imported


if __name__ == '__main__':
    import os
    from constants import DB_PATH

    excel = os.path.join(os.path.dirname(__file__), '..', 'agent_info.xlsx')
    db = Database(DB_PATH)
    count = import_from_excel(db, excel)
    print(f"Imported {count} agents")

    for a in db.get_all_agents():
        status = "ACTIVE" if a.is_active else "INACTIVE"
        print(f"  [{status}] {a.name}: {a.split_type}, "
              f"split={a.agent_split_pct}/{a.office_split_pct}, "
              f"cap={a.cap_amount}, fee={a.transaction_fee}, "
              f"contract={a.contract_date}")
    db.close()
